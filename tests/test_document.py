"""Document tool — a file they can open, Allow on write, no chat dump."""

from __future__ import annotations

from pathlib import Path

import pytest
from pypdf import PdfReader

from arelis.core.claims import detect_doc_ask, detect_document_ask
from arelis.core.plan_nudge import select_plan
from arelis.core.preflight import detect_intents
from arelis.core.skills import select_skill_ids
from arelis.tools import build_tool_registry
from arelis.tools.base import capability_class, confirm_args_blocked
from arelis.tools.confirm_copy import confirm_headline
from arelis.tools.document import DocumentTool
from arelis.workspace import WorkspaceRoots


@pytest.fixture
def project(tmp_path, monkeypatch):
    root = tmp_path / "project"
    root.mkdir()
    data = tmp_path / "appdata"
    data.mkdir()
    monkeypatch.setenv("ARELIS_DATA_DIR", str(data))
    workspace = WorkspaceRoots.from_config(
        {"workspace": {"roots": [{"name": "project", "path": str(root)}]}}
    )
    return root, data, workspace


@pytest.fixture
def tool(project):
    _root, _data, workspace = project
    return DocumentTool(workspace)


@pytest.mark.asyncio
async def test_writes_markdown(project, tool) -> None:
    _root, data, _ws = project
    result = await tool.run(
        format="md",
        title="Notes",
        body="Hello from Arelis.",
    )
    assert result.ok, result.output
    dest = Path(result.data["abs_path"])
    assert dest.suffix == ".md"
    text = dest.read_text(encoding="utf-8")
    assert "Notes" in text
    assert "Hello from Arelis." in text
    assert (data / "outputs" / "documents").is_dir()
    assert "open that file" in result.output.lower()


@pytest.mark.asyncio
async def test_writes_pdf_with_pages(project, tool) -> None:
    result = await tool.run(
        format="pdf",
        title="Dirac equation",
        body="# Dirac\n\nThe free Dirac equation is\n\n(i γ^μ ∂_μ − m) ψ = 0\n\n- spin 1/2\n- antimatter",
    )
    assert result.ok, result.output
    dest = Path(result.data["abs_path"])
    assert dest.suffix == ".pdf"
    n = len(PdfReader(str(dest)).pages)
    assert n >= 1
    assert "page" in result.output.lower()


@pytest.mark.asyncio
async def test_writes_docx(project, tool) -> None:
    result = await tool.run(
        format="docx",
        title="Memo",
        body="A short memo.",
    )
    assert result.ok, result.output
    dest = Path(result.data["abs_path"])
    assert dest.suffix == ".docx"
    assert dest.stat().st_size > 1000


@pytest.mark.asyncio
async def test_writes_xlsx_from_json_rows(project, tool) -> None:
    result = await tool.run(
        format="xlsx",
        title="Prices",
        rows='[["item","usd"],["alpha",344.72]]',
    )
    assert result.ok, result.output
    dest = Path(result.data["abs_path"])
    assert dest.suffix == ".xlsx"
    from openpyxl import load_workbook

    book = load_workbook(dest)
    assert book.active["A2"].value == "alpha"


@pytest.mark.asyncio
async def test_writes_csv(project, tool) -> None:
    result = await tool.run(
        format="csv",
        title="prices",
        body="item,usd\nalpha,344.72\n",
    )
    assert result.ok, result.output
    text = Path(result.data["abs_path"]).read_text(encoding="utf-8-sig")
    assert "alpha" in text


@pytest.mark.asyncio
async def test_filename_cannot_leave_documents_dir(project, tool) -> None:
    _root, data, _ws = project
    result = await tool.run(
        format="txt",
        title="x",
        body="hi",
        filename="..\\..\\escape.txt",
    )
    assert result.ok, result.output
    dest = Path(result.data["abs_path"]).resolve()
    folder = (data / "outputs" / "documents").resolve()
    assert dest.parent == folder


def test_empty_body_blocked_from_allow() -> None:
    assert confirm_args_blocked("document", {"format": "pdf"})


def test_document_needs_confirm_writes_and_skips_jobs(project) -> None:
    _root, _data, workspace = project
    config = {"tools": {}, "agent": {"confirm_writes": True}}
    registry = build_tool_registry(config, workspace, allow_send=True)
    args = {"format": "pdf", "title": "x", "body": "hello"}
    assert registry.get("document") is not None
    assert registry.needs_confirm("document", args)
    assert not registry.needs_confirm("document", args, confirm_writes=False)
    jobs = build_tool_registry(config, workspace, allow_send=False)
    assert jobs.get("document") is None
    assert capability_class("document") == "WRITE_LOCAL_ARTIFACT"
    assert confirm_headline("document", {"format": "pdf"}) == "write a pdf"
    assert confirm_headline("document", {"filename": "dirac.pdf"}) == "write dirac.pdf"
    assert (
        confirm_headline("document", {"filename": "dirac.pdf", "replace": "true"})
        == "replace dirac.pdf"
    )
    assert registry.get("document").rooms is config["_rooms"]
    detail = registry.describe_call(
        "document", {"format": "md", "title": "Notes", "body": "hi"}
    )
    assert "you can open" in detail
    assert "Lands in:" in detail


def test_create_pdf_is_not_a_pdf_read() -> None:
    ask = "create a pdf and i want a 5 page report in it about the physics behind the dirac equation"
    assert detect_document_ask(ask)
    assert not detect_doc_ask(ask)
    kinds = {h.kind for h in detect_intents(ask)}
    assert "document" in kinds
    assert "docs" not in kinds
    tools = {t for h in detect_intents(ask) for t in h.expected_tools}
    assert "document" in tools
    assert "doc_extract" not in tools
    plan = select_plan(ask)
    assert plan is not None and plan.id == "document"
    skills = select_skill_ids(
        ask, available_tools={"document", "doc_extract", "calculator"}
    )
    assert "document" in skills
    assert "docs" not in skills


def test_read_pdf_still_extracts() -> None:
    ask = "what does this pdf say about termination"
    assert detect_doc_ask(ask)
    assert not detect_document_ask(ask)
    plan = select_plan(ask)
    assert plan is not None and plan.id == "docs"


@pytest.mark.asyncio
async def test_room_with_folder_writes_to_documents(project) -> None:
    from arelis.rooms import RoomStore

    root, _data, workspace = project
    store = RoomStore(root / "rooms.yaml")
    store.update("physics", root="project", kind="writing")
    store.set_active("physics")
    tool = DocumentTool(workspace, store)
    result = await tool.run(format="md", title="Dirac", body="notes")
    assert result.ok, result.output
    dest = Path(result.data["abs_path"]).resolve()
    assert dest.parent == (root / "documents").resolve()
    assert "room" in result.output.lower()


@pytest.mark.asyncio
async def test_room_without_folder_uses_drop_tray(project) -> None:
    from arelis.rooms import RoomStore

    root, data, workspace = project
    store = RoomStore(root / "rooms.yaml")
    store.create("Scratch", kind="writing")
    store.set_active("scratch")
    tool = DocumentTool(workspace, store)
    result = await tool.run(format="txt", title="x", body="hi")
    assert result.ok, result.output
    dest = Path(result.data["abs_path"]).resolve()
    assert dest.parent == (data / "outputs" / "documents").resolve()
    assert "no folder" in result.output.lower()


@pytest.mark.asyncio
async def test_replace_overwrites_same_name(project, tool) -> None:
    first = await tool.run(format="txt", title="memo", body="one", filename="memo.txt")
    assert first.ok
    path = Path(first.data["abs_path"])
    second = await tool.run(
        format="txt",
        title="memo",
        body="two",
        filename="memo.txt",
        replace="true",
    )
    assert second.ok, second.output
    assert Path(second.data["abs_path"]).resolve() == path.resolve()
    assert path.read_text(encoding="utf-8").strip() == "two"
    assert second.data["replaced"] is True


@pytest.mark.asyncio
async def test_without_replace_keeps_both(project, tool) -> None:
    first = await tool.run(format="txt", title="memo", body="one", filename="memo.txt")
    second = await tool.run(format="txt", title="memo", body="two", filename="memo.txt")
    assert first.ok and second.ok
    assert Path(first.data["abs_path"]).resolve() != Path(second.data["abs_path"]).resolve()


@pytest.mark.asyncio
async def test_from_path_exports_markdown(project, tool) -> None:
    _root, data, _ws = project
    source = data / "outputs" / "documents" / "draft.md"
    source.parent.mkdir(parents=True, exist_ok=True)
    source.write_text("# Draft\n\nHello export.\n", encoding="utf-8")
    result = await tool.run(format="txt", from_path=str(source), filename="out.txt")
    assert result.ok, result.output
    text = Path(result.data["abs_path"]).read_text(encoding="utf-8")
    assert "Hello export." in text


@pytest.mark.asyncio
async def test_pdf_and_docx_render_markdown_tables(project, tool) -> None:
    body = "Intro\n\n| a | b |\n| --- | --- |\n| 1 | 2 |\n"
    pdf = await tool.run(format="pdf", title="Table", body=body)
    assert pdf.ok, pdf.output
    assert Path(pdf.data["abs_path"]).stat().st_size > 500
    docx = await tool.run(format="docx", title="Table", body=body)
    assert docx.ok, docx.output
    from docx import Document

    table = Document(docx.data["abs_path"]).tables[0]
    assert table.cell(0, 0).text == "a"
    assert table.cell(1, 1).text == "2"


def test_from_path_allows_empty_body() -> None:
    assert not confirm_args_blocked(
        "document", {"format": "pdf", "from_path": "notes.md"}
    )
    assert confirm_args_blocked("document", {"format": "pdf"})
